import openpyxl


def test_sheet_data(module_name):
    final = dict()
    work_book = openpyxl.load_workbook(
            "C:\\Users\\U6069615\\OneDrive - Clarivate Analytics\\Desktop\\PytestYokesh\\testdata.xlsx")
    sheet = work_book[module_name]
    for i in range(1, sheet.max_row):
        data_hashmap = dict()
        for j in range(2, sheet.max_column + 1):
            data_hashmap[sheet.cell(row=1, column=j).value] = sheet.cell(row=i + 1, column=j).value
        final[sheet.cell(row=i + 1, column=1).value] = data_hashmap
    return final


print(test_sheet_data("test_search")["test_verify_fielded_search"]["username"])



