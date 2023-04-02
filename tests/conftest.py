import json
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import pytest

driver = None


@pytest.fixture(autouse=True)
def setup(request, browser, environment):
    global driver
    components = request.node.nodeid.split("::")
    file_name = components[0].split("/")[-1]
    final_final_name = file_name.split('.')[0]
    # class_name = components[1]
    test_case_name = request.node.originalname
    test_data = test_sheet_data(final_final_name)
    request.cls.test_data = test_data[test_case_name]
    if browser == "chrome":
        chrome_options = Options()
        chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), chrome_options=chrome_options)
    elif browser == "firefox":
        driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))
    elif browser == "edge":
        driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))
    else:
        raise Exception("BROWSER UNAVAILABLE")
    request.cls.driver = driver
    if environment == "qa1":
        driver.get(get_environments(environment))
    elif environment == "qa2":
        driver.get(get_environments(environment))
    elif environment == "qa3":
        driver.get(get_environments(environment))
    elif environment == "qa4":
        driver.get(get_environments(environment))
    elif environment == "qa5":
        driver.get(get_environments(environment))
    elif environment == "automation":
        driver.get(get_environments(environment))
    else:
        raise Exception("NO SUITABLE ENVIRONMENT FOUND!!!")
    driver.maximize_window()
    time.sleep(3)
    yield driver
    driver.close()


def pytest_addoption(parser):
    parser.addoption("--browser")
    parser.addoption("--environment")


@pytest.fixture()
def browser(request):
    return request.config.getoption("--browser")


@pytest.fixture()
def environment(request):
    return request.config.getoption("--environment")


def test_sheet_data(module_name):
    final = dict()
    work_book = openpyxl.load_workbook(
            "C:\\Users\\U6069615\\OneDrive - Clarivate Analytics\\Desktop\\PytestYokesh\\testdata.xlsx")
    sheet = work_book[module_name]
    for i in range(1, sheet.max_row):
        data_hashmap = dict()
        for j in range(2, sheet.max_column + 1):
            data_hashmap[sheet.cell(row=1, column=j).value] = sheet.cell(row=i + 1, column=j).value
        final[sheet.cell(row=i + 1, column=1).value] = data_hashmap
    return final


def get_environments(environment):
    json_file_path = "C:\\Users\\U6069615\\OneDrive - Clarivate Analytics\\Desktop\\PytestYokesh\\configurations.json"
    f = open(json_file_path)
    data = json.load(f)
    return data['environments'][environment]


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])
    if report.when == 'call':
        xfail_state = hasattr(report, 'wasxfail')
        if (report.skipped and xfail_state) or (report.failed and not xfail_state):
            mydriver = item.funcargs['setup']
            screenshot = mydriver.get_screenshot_as_base64()
            extra.append(pytest_html.extras.image(screenshot, ''))
    report.extra = extra
